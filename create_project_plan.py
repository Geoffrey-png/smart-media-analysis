from datetime import date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference

out = r"E:\code\2026实训\智能传媒内容分析与推荐系统\智能传媒内容分析与推荐系统_14天项目计划.xlsx"
start = date(2026, 5, 29)
team = [
    ["王淦", "组长 / 全栈统筹", "项目计划、需求把控、架构设计、代码质量、关键模块联调", "全周期"],
    ["朱自豪", "后端与算法开发", "FastAPI接口、数据库模型、内容分析、推荐算法、权限与日志", "D1-D14"],
    ["秦梓洋", "前端与测试文档", "Vue页面、数据看板、交互联调、测试用例、演示材料", "D1-D14"],
]
milestones = [
    ["M1", "需求与架构确认", start, start + timedelta(days=1), "需求范围、数据库/接口草案、页面清单", "王淦", "未开始"],
    ["M2", "后端基础与核心数据闭环", start + timedelta(days=2), start + timedelta(days=4), "后端框架、模型、初始化数据、内容管理API", "朱自豪", "未开始"],
    ["M3", "内容分析与推荐核心能力", start + timedelta(days=5), start + timedelta(days=7), "内容分析、用户画像、推荐接口、审核流程", "朱自豪", "未开始"],
    ["M4", "前端页面与可视化闭环", start + timedelta(days=8), start + timedelta(days=10), "登录、仪表盘、内容/用户/推荐/审核页面", "秦梓洋", "未开始"],
    ["M5", "联调测试与演示交付", start + timedelta(days=11), start + timedelta(days=13), "联调修复、验收测试、README、演示材料", "王淦", "未开始"],
]
tasks = [
    [1, start, "需求复盘与范围冻结", "整理项目目标、角色权限、核心功能清单，明确14天交付边界", "王淦", "朱自豪、秦梓洋", "M1", "高", "需求清单、模块边界、风险清单", "未开始"],
    [2, start, "技术架构与数据库草案", "确认FastAPI+Vue3架构，梳理contents/users/behaviors/categories/tags等核心表", "王淦", "朱自豪", "M1", "高", "架构图说明、数据库字段草案", "未开始"],
    [3, start + timedelta(days=1), "接口清单与页面路由设计", "整理登录、内容、用户、行为、画像、推荐、审核、统计、上传等接口与前端页面路由", "王淦", "秦梓洋", "M1", "高", "接口清单、页面路由表", "未开始"],
    [4, start + timedelta(days=1), "开发环境与启动脚本检查", "检查后端虚拟环境、前端依赖、初始化脚本和README启动流程", "秦梓洋", "王淦", "M1", "中", "环境检查记录、可启动说明", "未开始"],
    [5, start + timedelta(days=2), "后端基础框架加固", "完善FastAPI入口、CORS、统一响应、异常处理、配置与数据库连接", "朱自豪", "王淦", "M2", "高", "后端基础框架可运行", "未开始"],
    [6, start + timedelta(days=2), "用户与权限模型实现", "完善JWT登录、角色权限、账号状态、管理员保护逻辑", "朱自豪", "王淦", "M2", "高", "登录认证与权限校验可用", "未开始"],
    [7, start + timedelta(days=3), "内容管理与文件上传API", "实现内容增删改查、分页筛选、状态管理、封面/附件上传接口", "朱自豪", "秦梓洋", "M2", "高", "内容管理API完成并可测试", "未开始"],
    [8, start + timedelta(days=3), "新闻采集与操作日志", "实现RSS新闻源采集、来源记录、采集审计日志", "朱自豪", "王淦", "M2", "高", "新闻采集闭环", "未开始"],
    [9, start + timedelta(days=4), "初始化演示数据完善", "补充内容、用户、行为、分类、标签和角色演示数据", "王淦", "朱自豪", "M2", "高", "演示数据库可复现", "未开始"],
    [10, start + timedelta(days=4), "后端接口自测", "用API文档或脚本测试核心接口，记录问题并修复", "王淦", "朱自豪", "M2", "高", "接口自测记录", "未开始"],
    [11, start + timedelta(days=5), "内容智能分析算法", "实现摘要、关键词、分类、情感、敏感词、热度分、质量分", "朱自豪", "王淦", "M3", "高", "内容分析接口可返回结果", "未开始"],
    [12, start + timedelta(days=6), "用户行为与画像更新", "实现view/like/favorite/comment/share/dislike行为记录及画像权重更新", "朱自豪", "秦梓洋", "M3", "高", "行为影响画像数据", "未开始"],
    [13, start + timedelta(days=6), "推荐算法与推荐接口", "实现热门推荐、画像匹配推荐、相似推荐、混合推荐和推荐原因", "朱自豪", "王淦", "M3", "高", "推荐接口可演示", "未开始"],
    [14, start + timedelta(days=7), "内容审核与统计接口", "实现待审核、通过、拒绝、下架、审核意见、推荐效果统计接口", "朱自豪", "秦梓洋", "M3", "高", "审核与统计接口可用", "未开始"],
    [15, start + timedelta(days=7), "算法和权限专项评审", "王淦审查算法合理性、权限边界和关键代码质量", "王淦", "朱自豪", "M3", "高", "评审问题清单和修复建议", "未开始"],
    [16, start + timedelta(days=8), "前端基础框架与布局", "完善Vue3、Pinia、路由、Axios封装、菜单权限和分析平台风格布局", "秦梓洋", "王淦", "M4", "高", "前端基础布局完成", "未开始"],
    [17, start + timedelta(days=8), "登录与管理员后台页面", "实现登录页、管理员后台、用户管理、角色状态展示", "秦梓洋", "朱自豪", "M4", "高", "登录和后台页面可操作", "未开始"],
    [18, start + timedelta(days=9), "内容资产与智能分析页面", "实现内容列表、详情、新增编辑、智能分析、审核操作页面", "秦梓洋", "朱自豪", "M4", "高", "内容业务页面闭环", "未开始"],
    [19, start + timedelta(days=9), "用户画像与推荐实验页面", "实现用户画像、推荐结果、推荐原因、推荐效果分析页面", "秦梓洋", "朱自豪", "M4", "高", "画像和推荐页面可演示", "未开始"],
    [20, start + timedelta(days=10), "数据看板与ECharts图表", "完成指标卡、行为趋势、分类分布、热门内容、CTR等图表", "秦梓洋", "王淦", "M4", "高", "数据看板完成", "未开始"],
    [21, start + timedelta(days=10), "前端交互与样式优化", "补充加载态、空状态、表单校验、删除确认、分析平台视觉优化", "秦梓洋", "王淦", "M4", "中", "前端体验优化记录", "未开始"],
    [22, start + timedelta(days=11), "前后端联调", "围绕登录、内容、采集、分析、画像、推荐、审核、统计进行全链路联调", "王淦", "朱自豪、秦梓洋", "M5", "高", "联调问题清单", "未开始"],
    [23, start + timedelta(days=11), "缺陷修复与回归测试", "修复联调发现的问题，完成关键路径回归测试", "朱自豪", "王淦、秦梓洋", "M5", "高", "缺陷修复记录、回归结果", "未开始"],
    [24, start + timedelta(days=12), "验收测试与演示脚本", "按验收标准检查启动、接口、页面、算法、推荐和数据看板，准备演示流程", "秦梓洋", "王淦", "M5", "高", "验收清单、演示脚本", "未开始"],
    [25, start + timedelta(days=12), "README与项目文档完善", "补充运行说明、账号说明、功能说明、扩展方向和已知问题", "王淦", "秦梓洋", "M5", "中", "README和项目文档", "未开始"],
    [26, start + timedelta(days=13), "最终代码质量检查", "检查目录结构、命名、异常处理、权限校验、注释、无明显语法错误", "王淦", "朱自豪", "M5", "高", "代码质量检查记录", "未开始"],
    [27, start + timedelta(days=13), "成果打包与答辩准备", "整理演示数据、截图、项目亮点、分工说明和风险说明", "王淦", "朱自豪、秦梓洋", "M5", "高", "可演示成果包", "未开始"],
]
risks = [
    ["R1", "14天周期较短，功能范围较大", "高", "高", "按P0/P1/P2分层交付，先保证可运行闭环", "王淦"],
    ["R2", "推荐算法效果可能偏演示级", "中", "中", "明确采用规则+轻量模型，保留后续大模型/NLP扩展点", "朱自豪"],
    ["R3", "前后端接口字段不一致导致联调延误", "中", "高", "D2完成接口清单，联调前冻结字段，变更需记录", "王淦"],
    ["R4", "真实RSS源不可用或格式不统一", "中", "中", "保留内置演示源和异常处理，采集失败不影响核心演示", "朱自豪"],
    ["R5", "页面像普通后台，分析平台感不足", "中", "中", "优先完成指标卡、图表、推荐效果分析和数据洞察首页", "秦梓洋"],
]

wb = Workbook()
ws = wb.active
ws.title = "项目概览"

blue = "1F4E78"
light_blue = "D9EAF7"
mid_blue = "5B9BD5"
light_green = "E2F0D9"
light_orange = "FCE4D6"
light_gray = "F2F2F2"
white = "FFFFFF"
dark = "1F1F1F"
border = Border(left=Side(style="thin", color="D9E1F2"), right=Side(style="thin", color="D9E1F2"), top=Side(style="thin", color="D9E1F2"), bottom=Side(style="thin", color="D9E1F2"))

for s in wb.worksheets:
    s.sheet_view.showGridLines = False

ws.merge_cells("A1:H1")
ws["A1"] = "智能传媒内容分析与推荐系统｜14天项目计划"
ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=18, color=white)
ws["A1"].fill = PatternFill("solid", fgColor=blue)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32
overview = [
    ["项目名称", "智能传媒内容分析与推荐系统", "计划周期", "2026-05-29 至 2026-06-11"],
    ["技术栈", "FastAPI + SQLAlchemy + SQLite + Vue 3 + Element Plus + ECharts", "团队规模", "3人"],
    ["组长", "王淦", "组员", "朱自豪、秦梓洋"],
    ["项目目标", "形成可运行、可演示、功能闭环完整的智能传媒内容分析与推荐系统", "交付重点", "内容分析、用户画像、个性化推荐、管理员后台、数据看板"],
]
for r_idx, row in enumerate(overview, 3):
    for c_idx, val in enumerate(row, 1):
        cell = ws.cell(r_idx, c_idx, val)
        cell.font = Font(name="Microsoft YaHei", bold=c_idx in [1,3], color=dark)
        cell.fill = PatternFill("solid", fgColor=light_blue if c_idx in [1,3] else white)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
ws.merge_cells("B6:B6")
ws.merge_cells("D6:H6")
for col in range(1,9):
    ws.column_dimensions[get_column_letter(col)].width = [14,32,14,30,16,16,16,16][col-1]

ws["A8"] = "核心交付范围"
ws["A8"].font = Font(name="Microsoft YaHei", bold=True, size=13, color=blue)
deliverables = ["登录认证与角色权限", "新闻采集与内容资产管理", "内容摘要/关键词/分类/情感/敏感词分析", "用户行为记录与用户画像", "热门/个性化/相似/混合推荐", "内容审核与操作审计", "管理员后台与数据看板", "联调测试、README和演示材料"]
for i, item in enumerate(deliverables, 9):
    ws.cell(i, 1, i-8)
    ws.cell(i, 2, item)
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
    for c in range(1,9):
        ws.cell(i,c).border = border
        ws.cell(i,c).alignment = Alignment(vertical="center")
    ws.cell(i,1).fill = PatternFill("solid", fgColor=light_gray)
    ws.cell(i,2).font = Font(name="Microsoft YaHei")

ws["A18"] = "计划说明"
ws["A18"].font = Font(name="Microsoft YaHei", bold=True, size=13, color=blue)
notes = [
    "1. 计划从明天（2026-05-29）开始，连续14天推进，默认包含周末用于实训冲刺。",
    "2. 分工采用主责+协作机制：主责人对交付质量负责，协作人负责联调、评审或补位。",
    "3. 第1阶段优先保证可运行闭环；算法以规则和轻量模型为主，保留后续扩展接口。",
    "4. 王淦作为组长负责范围冻结、架构评审、联调推进、代码质量与最终交付把关。",
]
for i, n in enumerate(notes, 19):
    ws.cell(i,1,n)
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    ws.cell(i,1).alignment = Alignment(wrap_text=True)
    ws.cell(i,1).font = Font(name="Microsoft YaHei")

ms = wb.create_sheet("里程碑")
ms.append(["编号", "里程碑", "开始日期", "结束日期", "交付物", "负责人", "状态"])
for row in milestones:
    ms.append(row)

plan = wb.create_sheet("14天时间表")
headers = ["序号", "日期", "星期", "工作内容", "任务说明", "主责人", "协作人", "里程碑", "优先级", "交付物", "状态"]
plan.append(headers)
weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
for row in tasks:
    d = row[1]
    plan.append([row[0], d, weekdays[d.weekday()], *row[2:]])

people = wb.create_sheet("人员分配")
people.append(["姓名", "角色", "主要职责", "参与周期"])
for row in team:
    people.append(row)
people.append([])
people.append(["姓名", "负责任务数", "高优先级任务数", "涉及里程碑", "备注"])
for idx, name in enumerate(["王淦", "朱自豪", "秦梓洋"], 7):
    people.cell(idx, 1, name)
    people.cell(idx, 2, f'=COUNTIF(\'14天时间表\'!F:F,A{idx})')
    people.cell(idx, 3, f'=COUNTIFS(\'14天时间表\'!F:F,A{idx},\'14天时间表\'!I:I,"高")')
    people.cell(idx, 4, f'=TEXTJOIN("、",TRUE,UNIQUE(FILTER(\'14天时间表\'!H:H,\'14天时间表\'!F:F=A{idx})))')
    people.cell(idx, 5, "按主责任务统计")

risk = wb.create_sheet("风险与验收")
risk.append(["风险编号", "风险描述", "概率", "影响", "应对策略", "负责人"])
for row in risks:
    risk.append(row)
risk.append([])
risk.append(["验收项", "验收标准", "负责人", "状态"])
acceptance = [
    ["后端服务", "FastAPI可启动，/docs可访问，核心接口返回统一格式", "朱自豪", "未开始"],
    ["前端服务", "Vue页面可启动，核心页面可访问，菜单权限显示合理", "秦梓洋", "未开始"],
    ["内容闭环", "内容可采集、新增、编辑、分析、审核和展示", "朱自豪", "未开始"],
    ["推荐闭环", "行为记录可影响画像，推荐结果包含推荐原因", "朱自豪", "未开始"],
    ["数据看板", "指标卡、趋势图、分类分布、推荐效果统计可展示", "秦梓洋", "未开始"],
    ["文档交付", "README、演示账号、启动说明、项目亮点和已知问题完整", "王淦", "未开始"],
]
for row in acceptance:
    risk.append(row)

for sheet in [ms, plan, people, risk]:
    sheet.sheet_view.showGridLines = False
    max_col = sheet.max_column
    for cell in sheet[1]:
        cell.font = Font(name="Microsoft YaHei", bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10, bold=cell.font.bold, color=cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else dark)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(1, max_col + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 16
    sheet.freeze_panes = "A2"

for col, width in enumerate([8,14,10,24,46,12,20,12,10,28,10],1):
    plan.column_dimensions[get_column_letter(col)].width = width
for col, width in enumerate([10,22,14,14,40,14,12],1):
    ms.column_dimensions[get_column_letter(col)].width = width
for col, width in enumerate([12,20,46,18,14],1):
    people.column_dimensions[get_column_letter(col)].width = width
for col, width in enumerate([12,34,10,10,48,14],1):
    risk.column_dimensions[get_column_letter(col)].width = width

for sheet, ref, name in [(ms, "A1:G6", "MilestoneTable"), (plan, "A1:K28", "ScheduleTable"), (people, "A1:D4", "PeopleTable"), (risk, "A1:F6", "RiskTable")]:
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

for row in plan.iter_rows(min_row=2, max_row=plan.max_row):
    if row[8].value == "高":
        row[8].fill = PatternFill("solid", fgColor=light_orange)
    if row[10].value == "未开始":
        row[10].fill = PatternFill("solid", fgColor=light_gray)

chart_sheet = wb.create_sheet("任务分布图")
chart_sheet.sheet_view.showGridLines = False
chart_sheet.append(["成员", "负责任务数", "高优先级任务数"])
for name in ["王淦", "朱自豪", "秦梓洋"]:
    chart_sheet.append([name, f'=COUNTIF(\'14天时间表\'!F:F,A{chart_sheet.max_row+1})', f'=COUNTIFS(\'14天时间表\'!F:F,A{chart_sheet.max_row+1},\'14天时间表\'!I:I,"高")'])
for cell in chart_sheet[1]:
    cell.font = Font(name="Microsoft YaHei", bold=True, color=white)
    cell.fill = PatternFill("solid", fgColor=blue)
for col in range(1,4):
    chart_sheet.column_dimensions[get_column_letter(col)].width = 18
chart = BarChart()
chart.title = "团队任务分布"
chart.y_axis.title = "任务数"
chart.x_axis.title = "成员"
data = Reference(chart_sheet, min_col=2, max_col=3, min_row=1, max_row=4)
cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=4)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 8
chart.width = 16
chart_sheet.add_chart(chart, "E2")

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=cell.font.sz or 10, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else dark)

wb.save(out)

check = load_workbook(out, data_only=False)
errors = []
for ws in check.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("#"):
                errors.append(f"{ws.title}!{cell.coordinate}:{cell.value}")
if errors:
    raise SystemExit("Formula/value errors found: " + "; ".join(errors))
print(out)
